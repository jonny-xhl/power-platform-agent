"""
Dataverse Entity Model Design Template Generator

Generates a comprehensive Excel template for designing Microsoft Dataverse
entity models with support for:
- Entity/field definitions (Dataverse compatible)
- Multi-tab form layouts
- View configurations
- Option sets/choices
- Business rules
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
import sys
import os

def create_template(output_file='ba_system_design_template.xlsx'):
    """Generate the Dataverse entity model design template"""

    wb = Workbook()

    # Styles
    header_fill = PatternFill('solid', start_color='4472C4', end_color='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    tab_fill = PatternFill('solid', start_color='ED7D31', end_color='ED7D31')
    tab_font = Font(bold=True, color='FFFFFF')
    section_fill = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # ============ Sheet 1: 使用说明 ============
    index_sheet = wb.active
    index_sheet.title = '01_使用说明'
    index_sheet.column_dimensions['A'].width = 25
    index_sheet.column_dimensions['B'].width = 60

    index_sheet['A1'] = 'BA系统设计模板'
    index_sheet['A1'].font = Font(bold=True, size=16)
    index_sheet['A2'] = '版本: 2.1 - Dataverse兼容 (支持多Tab表单)'
    index_sheet['A2'].font = Font(bold=True, size=10)
    index_sheet['A3'] = '本模板用于定义系统实体模型、视图和表单设计（兼容Microsoft Dataverse）'
    index_sheet['A3'].alignment = left_align
    index_sheet.merge_cells('A3:B3')

    index_sheet['A4'] = '重要提示:'
    index_sheet['A4'].font = Font(bold=True, color='C00000')

    important_notes = [
        ['说明项', '详细说明'],
        ['主字段', '每个实体必须有且仅有一个主字段(Primary Name)，用于记录标题显示，通常为Text类型'],
        ['多Tab表单', '表单支持多个Tab页，每个Tab可包含多个分组(Section)'],
        ['数据类型', '请参考"07_Dataverse类型"工作表中的官方数据类型'],
    ]

    for row_idx, row_data in enumerate(important_notes, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = index_sheet.cell(row_idx, col_idx, value)
            if row_idx == 5:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            cell.border = border

    instructions = [
        ['', ''],
        ['工作表', '说明'],
        ['01_使用说明', '模板使用指南和说明'],
        ['02_实体模型', '定义数据实体、字段、类型、关系（Dataverse兼容）'],
        ['03_视图定义', '定义列表视图、表格视图的列配置和显示规则'],
        ['04_表单设计', '多Tab表单布局设计（可视化画布）'],
        ['05_枚举选项集', '定义下拉选项、状态值等枚举类型'],
        ['06_业务规则', '定义业务规则、计算字段、触发器等'],
        ['07_Dataverse类型', 'Dataverse数据类型参考说明'],
    ]

    for row_idx, row_data in enumerate(instructions, 10):
        for col_idx, value in enumerate(row_data, 1):
            cell = index_sheet.cell(row_idx, col_idx, value)
            if row_idx == 10:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            cell.border = border

    # ============ Sheet 2: 实体模型 ============
    entity_sheet = wb.create_sheet('02_实体模型')
    cols = 'A,B,C,D,E,F,G,H,I,J,K,L,M,N,O'.split(',')
    widths = [18, 5, 20, 18, 20, 18, 10, 20, 10, 10, 12, 12, 15, 15, 12]
    for c, w in zip(cols, widths):
        entity_sheet.column_dimensions[c].width = w

    entity_headers = [
        '实体名称', '序号', '字段名称', '显示名称', '英文显示名称', '数据类型',
        '长度', '默认值', '是否必填', '是否唯一', '是否主字段', '可搜索',
        '关联实体', '关联类型', '说明'
    ]

    for col_idx, header in enumerate(entity_headers, 1):
        cell = entity_sheet.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Sample account entity
    example_entity = [
        ['account', '1', 'account_name', '客户名称', 'Account Name', 'Text', '200', '', '是', '是', '是', '是', '', '', '主字段-必须且唯一'],
        ['', '2', 'account_id', '客户编号', 'Account ID', 'Autonumber', '', '', '是', '是', '否', '是', '', '', '自动编号'],
        ['', '3', 'account_type', '客户类型', 'Account Type', 'Choice', '', '1', '是', '否', '否', '是', '', '', '选项集'],
        ['', '4', 'telephone1', '联系电话', 'Main Phone', 'Phone', '20', '', '否', '否', '否', '是', '', '', ''],
        ['', '5', 'emailaddress1', '电子邮箱', 'Email', 'Email', '100', '', '否', '否', '否', '是', '', '', ''],
        ['', '6', 'statuscode', '状态', 'Status', 'Choice', '', '1', '是', '否', '否', '是', '', '', '状态原因'],
        ['', '7', 'createdon', '创建日期', 'Created On', 'Date and Time', '', '', '否', '否', '否', '是', '', '', '系统字段'],
        ['', '8', 'primarycontactid', '主联系人', 'Primary Contact', 'Lookup', '', '', '否', '否', '否', '是', 'contact', 'Many-to-One', ''],
        ['', '9', 'ownerid', '所有者', 'Owner', 'Owner', '', '', '是', '否', '否', '是', 'systemuser', 'Many-to-One', '系统字段'],
    ]

    for row_idx, row_data in enumerate(example_entity, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = entity_sheet.cell(row_idx, col_idx, value)
            if value == '是':
                cell.font = Font(color='008000')
            cell.border = border
            cell.alignment = left_align

    # Sample contact entity
    example_contact = [
        ['contact', '1', 'fullname', '全名', 'Full Name', 'Text', '200', '', '是', '是', '是', '是', '', '', '主字段-必须且唯一'],
        ['', '2', 'firstname', '名', 'First Name', 'Text', '100', '', '是', '否', '否', '是', '', '', ''],
        ['', '3', 'lastname', '姓', 'Last Name', 'Text', '100', '', '是', '否', '否', '是', '', '', ''],
        ['', '4', 'emailaddress1', '电子邮箱', 'Email', 'Email', '100', '', '否', '否', '否', '是', '', '', ''],
        ['', '5', 'telephone1', '联系电话', 'Main Phone', 'Phone', '20', '', '否', '否', '否', '是', '', '', ''],
        ['', '6', 'parentcustomerid', '所属客户', 'Parent Account', 'Customer', '', '', '否', '否', '否', '是', 'account', 'Many-to-One', '可关联Account或Contact'],
    ]

    for row_idx, row_data in enumerate(example_contact, 12):
        for col_idx, value in enumerate(row_data, 1):
            cell = entity_sheet.cell(row_idx, col_idx, value)
            if value == '是':
                cell.font = Font(color='008000')
            cell.border = border
            cell.alignment = left_align

    # ============ Sheet 3: 视图定义 ============
    view_sheet = wb.create_sheet('03_视图定义')
    view_widths = [18, 18, 5, 20, 18, 8, 10, 10, 10, 10, 12, 30]
    for c, w in zip('A,B,C,D,E,F,G,H,I,J,K,L'.split(','), view_widths):
        view_sheet.column_dimensions[c].width = w

    view_headers = [
        '视图名称', '实体名称', '序号', '字段名称', '显示名称',
        '列宽', '对齐方式', '是否可排', '是否可筛', '排序方式',
        '聚合方式', '格式化规则'
    ]

    for col_idx, header in enumerate(view_headers, 1):
        cell = view_sheet.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    example_view = [
        ['Active Accounts', 'account', '1', 'account_name', '客户名称', '200', 'left', '是', '是', 'asc', '', ''],
        ['', '', '2', 'account_type', '客户类型', '100', 'center', '是', '是', '', '', ''],
        ['', '', '3', 'telephone1', '联系电话', '120', 'left', '是', '是', '', '', ''],
        ['', '', '4', 'emailaddress1', '电子邮箱', '180', 'left', '是', '是', '', '', ''],
        ['', '', '5', 'statuscode', '状态', '80', 'center', '是', '是', '', '', '1=激活(绿),2=禁用(红)'],
        ['', '', '6', 'createdon', '创建日期', '120', 'center', '是', '是', 'desc', '', 'yyyy-MM-dd'],
    ]

    for row_idx, row_data in enumerate(example_view, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = view_sheet.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = left_align

    # ============ Sheet 4: 表单设计 (多Tab可视化布局) ============
    form_sheet = wb.create_sheet('04_表单设计')

    for col in 'ABCDEFGHI':
        form_sheet.column_dimensions[col].width = 18

    for row in range(1, 50):
        form_sheet.row_dimensions[row].height = 25

    # Header
    form_sheet.merge_cells('A1:I1')
    title_cell = form_sheet['A1']
    title_cell.value = '表单设计画布 (支持多Tab)'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align
    title_cell.fill = PatternFill('solid', start_color='E7E6E6', end_color='E7E6E6')

    form_sheet['A2'] = '表单名称:'
    form_sheet['B2'] = 'Account Main Form'
    form_sheet['D2'] = '实体名称:'
    form_sheet['E2'] = 'account'

    # Tab 1: 常规
    form_sheet.merge_cells('A4:I4')
    tab1_cell = form_sheet['A4']
    tab1_cell.value = '【Tab 1: 常规】'
    tab1_cell.font = tab_font
    tab1_cell.fill = tab_fill
    tab1_cell.alignment = center_align

    # Section: 基本信息
    form_sheet.merge_cells('A5:I6')
    section1 = form_sheet['A5']
    section1.value = '分组: 基本信息'
    section1.font = Font(bold=True)
    section1.fill = section_fill
    section1.alignment = Alignment(horizontal='left', vertical='center')

    tab1_fields = [
        ('A7', 'account_name', '客户名称 *'),
        ('E7', 'account_id', '客户编号 *'),
        ('A8', 'account_type', '客户类型 *'),
        ('E8', 'statuscode', '状态 *'),
    ]

    for cell_pos, field_name, label in tab1_fields:
        label_cell = form_sheet[cell_pos]
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        label_cell.comment = Comment(field_name, '字段名称')

    # Section: 联系信息
    form_sheet.merge_cells('A9:I10')
    section2 = form_sheet['A9']
    section2.value = '分组: 联系信息'
    section2.font = Font(bold=True)
    section2.fill = section_fill
    section2.alignment = Alignment(horizontal='left', vertical='center')

    contact_fields = [
        ('A11', 'telephone1', '联系电话'),
        ('E11', 'emailaddress1', '电子邮箱'),
        ('A12', 'websiteurl', '网站'),
        ('E12', 'fax', '传真'),
    ]

    for cell_pos, field_name, label in contact_fields:
        label_cell = form_sheet[cell_pos]
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        label_cell.comment = Comment(field_name, '字段名称')

    # Tab 2: 详细信息
    form_sheet.merge_cells('A14:I14')
    tab2_cell = form_sheet['A14']
    tab2_cell.value = '【Tab 2: 详细信息】'
    tab2_cell.font = tab_font
    tab2_cell.fill = tab_fill
    tab2_cell.alignment = center_align

    form_sheet.merge_cells('A15:I16')
    section3 = form_sheet['A15']
    section3.value = '分组: 公司信息'
    section3.font = Font(bold=True)
    section3.fill = section_fill
    section3.alignment = Alignment(horizontal='left', vertical='center')

    detail_fields = [
        ('A17', 'industrycode', '行业'),
        ('E17', 'numberofemployees', '员工数'),
        ('A18', 'revenue', '年收入'),
        ('E18', 'tickersymbol', '股票代码'),
    ]

    for cell_pos, field_name, label in detail_fields:
        label_cell = form_sheet[cell_pos]
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        label_cell.comment = Comment(field_name, '字段名称')

    # Tab 3: 关系
    form_sheet.merge_cells('A24:I24')
    tab3_cell = form_sheet['A24']
    tab3_cell.value = '【Tab 3: 关系】'
    tab3_cell.font = tab_font
    tab3_cell.fill = tab_fill
    tab3_cell.alignment = center_align

    form_sheet.merge_cells('A25:I26')
    section5 = form_sheet['A25']
    section5.value = '分组: 主要关系'
    section5.font = Font(bold=True)
    section5.fill = section_fill
    section5.alignment = Alignment(horizontal='left', vertical='center')

    relation_fields = [
        ('A27', 'primarycontactid', '主联系人'),
        ('E27', 'parentaccountid', '上级客户'),
    ]

    for cell_pos, field_name, label in relation_fields:
        label_cell = form_sheet[cell_pos]
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        label_cell.comment = Comment(field_name, '字段名称')

    # Field properties config
    for col, width in [('K',5), ('L',18), ('M',12), ('N',12), ('O',12), ('P',30), ('Q',30)]:
        form_sheet.column_dimensions[col].width = width

    form_sheet.merge_cells('L1:Q1')
    config_title = form_sheet['L1']
    config_title.value = '字段属性配置'
    config_title.font = Font(bold=True, size=12)
    config_title.alignment = center_align
    config_title.fill = PatternFill('solid', start_color='E7E6E6', end_color='E7E6E6')

    field_config_headers = ['字段名称', '是否只读', '是否必填', '验证规则', '提示信息', 'Tab位置']
    for col_idx, header in enumerate(field_config_headers, 12):
        cell = form_sheet.cell(2, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    field_config_data = [
        ['account_name', '否', '是', '长度2-200字符', '请输入客户名称', 'Tab1-基本信息'],
        ['account_id', '是', '是', '', '自动编号', 'Tab1-基本信息'],
        ['account_type', '否', '是', '', '选择客户类型', 'Tab1-基本信息'],
        ['statuscode', '否', '是', '', '选择状态', 'Tab1-基本信息'],
        ['telephone1', '否', '否', '手机号格式', '请输入联系电话', 'Tab1-联系信息'],
        ['emailaddress1', '否', '否', '邮箱格式', '请输入电子邮箱', 'Tab1-联系信息'],
        ['websiteurl', '否', '否', 'URL格式', '请输入网站地址', 'Tab1-联系信息'],
        ['fax', '否', '否', '', '请输入传真号码', 'Tab1-联系信息'],
        ['industrycode', '否', '否', '', '选择行业', 'Tab2-公司信息'],
        ['numberofemployees', '否', '否', '正整数', '请输入员工数', 'Tab2-公司信息'],
        ['revenue', '否', '否', '货币格式', '请输入年收入', 'Tab2-公司信息'],
        ['tickersymbol', '否', '否', '', '请输入股票代码', 'Tab2-公司信息'],
        ['primarycontactid', '否', '否', '', '选择主联系人', 'Tab3-主要关系'],
        ['parentaccountid', '否', '否', '', '选择上级客户', 'Tab3-主要关系'],
    ]

    for row_idx, row_data in enumerate(field_config_data, 3):
        for col_idx, value in enumerate(row_data, 12):
            cell = form_sheet.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = left_align

    # ============ Sheet 5: 枚举选项集 ============
    enum_sheet = wb.create_sheet('05_枚举选项集')
    enum_cols = 'A,B,C,D,E,F,G'.split(',')
    enum_widths = [18, 10, 20, 25, 25, 8, 12]
    for c, w in zip(enum_cols, enum_widths):
        enum_sheet.column_dimensions[c].width = w

    enum_headers = ['选项集名称', '序号', '选项值', '显示名称', '英文显示名称', '排序', '颜色标记']
    for col_idx, header in enumerate(enum_headers, 1):
        cell = enum_sheet.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    example_enum = [
        ['account_type', '1', '1', '个人客户', 'Individual', '1', '#2E75B5'],
        ['', '2', '2', '企业客户', 'Enterprise', '2', '#ED7D31'],
        ['', '3', '3', '政府机构', 'Government', '3', '#70AD47'],
        ['statuscode', '1', '1', '激活', 'Active', '1', '#70AD47'],
        ['', '2', '2', '禁用', 'Inactive', '2', '#C00000'],
        ['', '3', '3', '待审核', 'Pending', '3', '#FFC000'],
    ]

    for row_idx, row_data in enumerate(example_enum, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = enum_sheet.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = left_align

    # ============ Sheet 6: 业务规则 ============
    rule_sheet = wb.create_sheet('06_业务规则')
    for col, width in zip('A,B,C,D,E,F,G'.split(','), [18, 15, 35, 35, 15, 10, 10]):
        rule_sheet.column_dimensions[col].width = width

    rule_headers = ['规则名称', '实体名称', '触发条件', '执行动作', '执行时机', '优先级', '状态']
    for col_idx, header in enumerate(rule_headers, 1):
        cell = rule_sheet.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    example_rule = [
        ['客户名称唯一性', 'account', 'account_name发生变化', '检查客户名称是否已存在', '保存前', '1', '启用'],
        ['客户邮箱格式', 'account', 'emailaddress1字段不为空', '验证邮箱格式是否正确', '保存前', '2', '启用'],
    ]

    for row_idx, row_data in enumerate(example_rule, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = rule_sheet.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = left_align

    # ============ Sheet 7: Dataverse类型参考 ============
    type_sheet = wb.create_sheet('07_Dataverse类型')
    for col, width in zip('A,B,C,D,E,F'.split(','), [25, 25, 50, 15, 30, 15]):
        type_sheet.column_dimensions[col].width = width

    type_headers = ['数据类型', 'API类型', '说明', '最大长度', '是否可用作主字段', '示例']
    for col_idx, header in enumerate(type_headers, 1):
        cell = type_sheet.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    dataverse_types = [
        ['Text', 'StringType', '单行文本，最常用的字段类型', '4000', '是', 'account_name'],
        ['Multiline Text', 'MemoType', '多行文本，用于长文本内容', '1048576', '否', 'description'],
        ['Email', 'StringType(Email)', '电子邮件地址', '4000', '否', 'emailaddress1'],
        ['Phone', 'StringType(Phone)', '电话号码', '200', '否', 'telephone1'],
        ['URL', 'StringType(URL)', '网址链接', '4000', '否', 'websiteurl'],
        ['Date and Time', 'DateTimeType', '日期和时间', '', '否', 'createdon'],
        ['Whole Number', 'IntegerType', '整数类型', '', '否', 'numberofemployees'],
        ['Decimal Number', 'DecimalType', '十进制数，精度10位', '', '否', 'price'],
        ['Currency', 'MoneyType', '货币类型', '', '否', 'revenue'],
        ['Floating Point Number', 'DoubleType', '浮点数，精度5位', '', '否', 'temperature'],
        ['Yes/No', 'BooleanType', '布尔值(是/否)', '', '否', 'donotemail'],
        ['Choice', 'PicklistType', '选项集(单选)', '', '否', 'account_type'],
        ['MultiSelect Choice', 'MultiSelectPicklistType', '选项集(多选)', '', '否', 'interests'],
        ['Lookup', 'LookupType', '查找字段(多对一关系)', '', '否', 'primarycontactid'],
        ['Customer', 'LookupType(Customer)', '客户查找(可关联Account或Contact)', '', '否', 'parentcustomerid'],
        ['Owner', 'LookupType(Owner)', '所有者(可关联User或Team)', '', '否', 'ownerid'],
        ['Unique Identifier', 'UniqueidentifierType', '唯一标识符(GUID)，系统自动生成', '', '否', '{xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx}'],
        ['Image', 'ImageType', '图像字段，每个表仅可有一个', '', '否', 'entityimage'],
        ['File', 'FileType', '文件附件', '', '否', 'attachment'],
        ['Autonumber', 'StringType(Autonumber)', '自动编号，系统自动生成', '', '否', 'account_number'],
    ]

    for row_idx, row_data in enumerate(dataverse_types, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = type_sheet.cell(row_idx, col_idx, value)
            if value == '是':
                cell.font = Font(color='008000', bold=True)
            elif value == '否':
                cell.font = Font(color='C00000')
            cell.border = border
            cell.alignment = left_align

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save
    wb.save(output_file)
    print(f'Template created: {output_file}')
    print('Version: 2.1 - Dataverse Compatible (Multi-Tab Form Support)')
    return output_file


def main():
    """Main entry point"""
    if len(sys.argv) > 1:
        output_file = sys.argv[1]
    else:
        output_file = 'ba_system_design_template.xlsx'

    create_template(output_file)


if __name__ == '__main__':
    main()
