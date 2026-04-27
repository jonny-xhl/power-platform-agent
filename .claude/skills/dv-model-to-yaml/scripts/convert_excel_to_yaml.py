"""
Excel实体设计转YAML元数据转换器

将Excel中的Dataverse实体设计转换为符合Schema规范的YAML元数据文件
支持根据naming_rules.yaml配置自动转换schema name
"""

from openpyxl import load_workbook
import yaml
import re
from pathlib import Path
from typing import Dict, List, Any, Optional


class NamingRuleEngine:
    """命名规则引擎"""

    def __init__(self, config_path: str = None):
        """加载命名规则配置"""
        self.config = {
            'prefix': 'new',
            'schema_name': {
                'style': 'lowercase',
                'separator': '_',
                'auto_prefix': True
            },
            'standard_entities': [],
            'transformations': []
        }

        if config_path and Path(config_path).exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
                if 'naming' in config:
                    self.config.update(config['naming'])

    def convert_schema_name(self, input_name: str) -> str:
        """
        根据命名规则转换schema name

        Args:
            input_name: 输入的表名

        Returns:
            转换后的schema name
        """
        # 检查是否为标准实体
        if input_name in self.config.get('standard_entities', []):
            return input_name

        prefix = self.config['prefix']
        style = self.config['schema_name']['style']
        separator = self.config['schema_name']['separator']
        auto_prefix = self.config['schema_name']['auto_prefix']

        result = input_name

        # 应用转换规则
        # 1. 在小写和大写字母之间插入下划线 (AccountNumber → Account_Number)
        result = re.sub(r'([a-z])([A-Z])', r'\1\2', result)

        # 2. 根据目标风格转换
        if style == 'lowercase':
            result = result.lower()
            if separator:
                # 替换驼峰为分隔符
                result = re.sub(r'([a-z0-9])([A-Z])', rf'\1{separator}\2', result)
                result = result.lower()
        elif style == 'camelCase':
            # 移除下划线并转为驼峰
            parts = result.split('_')
            result = parts[0] + ''.join(p.capitalize() for p in parts[1:])
        elif style == 'PascalCase':
            # 转为大驼峰
            parts = result.split('_')
            result = ''.join(p.capitalize() for p in parts)

        # 添加前缀
        if auto_prefix and prefix:
            if not result.startswith(prefix + separator) and not result.startswith(prefix):
                if style == 'lowercase':
                    result = f"{prefix}{separator}{result}"
                else:
                    result = f"{prefix}{result}"

        return result

    def convert_attribute_name(self, input_name: str) -> str:
        """
        转换字段名

        Args:
            input_name: 输入的字段名

        Returns:
            转换后的字段名
        """
        prefix = self.config['prefix']
        style = self.config['schema_name']['style']
        separator = self.config['schema_name']['separator']

        # 字段通常也添加前缀
        result = input_name

        if style == 'lowercase':
            # 转为小写加下划线
            result = re.sub(r'([a-z])([A-Z])', rf'\1{separator}\2', result).lower()
            if prefix:
                result = f"{prefix}{separator}{result}"
        elif style == 'camelCase':
            # 转为驼峰
            parts = result.split('_')
            result = parts[0] + ''.join(p.capitalize() for p in parts[1:])
            if prefix:
                result = f"{prefix}{result.capitalize()}"
        elif style == 'PascalCase':
            # 转为大驼峰
            parts = result.split('_')
            result = ''.join(p.capitalize() for p in parts)
            if prefix:
                result = f"{prefix.capitalize()}{result}"

        return result


class ExcelToYamlConverter:
    """Excel到YAML转换器"""

    # Excel数据类型到YAML类型映射
    TYPE_MAPPING = {
        'Text': 'String',
        'Multiline Text': 'Memo',
        'Email': 'String',
        'Phone': 'String',
        'URL': 'String',
        'Date and Time': 'DateTime',
        'Whole Number': 'Integer',
        'Decimal Number': 'Decimal',
        'Currency': 'Money',
        'Floating Point Number': 'Double',
        'Yes/No': 'Boolean',
        'Choice': 'Picklist',
        'MultiSelect Choice': 'MultiSelectPicklist',
        'Lookup': 'Lookup',
        'Customer': 'Customer',
        'Owner': 'Owner',
        'Autonumber': 'String',
        'Unique Identifier': 'Uniqueidentifier'
    }

    def __init__(self, excel_path: str, config_path: str = None):
        """
        初始化转换器

        Args:
            excel_path: Excel文件路径
            config_path: 命名规则配置文件路径
        """
        self.excel_path = excel_path
        self.naming_engine = NamingRuleEngine(config_path)
        self.tables = {}
        self.optionsets = {}
        self.forms = {}
        self.views = {}

    def load_excel(self):
        """加载Excel文件"""
        try:
            wb = load_workbook(self.excel_path, read_only=True)

            # 读取实体模型工作表
            if '02_实体模型' in wb.sheetnames:
                self._parse_entity_model(wb['02_实体模型'])

            # 读取选项集工作表
            if '05_枚举选项集' in wb.sheetnames:
                self._parse_optionsets(wb['05_枚举选项集'])

            # 读取表单设计工作表
            if '04_表单设计' in wb.sheetnames:
                self._parse_forms(wb['04_表单设计'])

            # 读取视图定义工作表
            if '03_视图定义' in wb.sheetnames:
                self._parse_views(wb['03_视图定义'])

            wb.close()
            return True, "Excel加载成功"
        except Exception as e:
            return False, f"Excel加载失败: {str(e)}"

    def _parse_entity_model(self, sheet):
        """解析实体模型工作表"""
        current_entity = None
        headers = []

        # 读取表头
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [cell or '' for cell in row]
            break

        # 查找列索引
        col_idx = {name: idx for idx, name in enumerate(headers)}

        # 解析数据行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            entity_name = str(row[col_idx.get('实体名称', 1)] or '').strip()

            # 新实体开始
            if entity_name and entity_name != '':
                current_entity = entity_name
                # 应用命名规则转换schema name
                schema_name = self.naming_engine.convert_schema_name(entity_name)

                if current_entity not in self.tables:
                    self.tables[current_entity] = {
                        'original_name': current_entity,
                        'schema_name': schema_name,
                        'display_name': '',
                        'description': '',
                        'ownership_type': 'UserOwned',
                        'attributes': [],
                        'relationships': []
                    }

            # 解析字段
            if current_entity:
                field_name = str(row[col_idx.get('字段名称', 2)] or '').strip()
                if field_name and field_name != '':
                    attr = self._parse_attribute(row, col_idx)
                    self.tables[current_entity]['attributes'].append(attr)

                    # 如果是第一个字段，作为表的显示名称
                    if len(self.tables[current_entity]['attributes']) == 1:
                        display_name = str(row[col_idx.get('显示名称', 3)] or '').strip()
                        if display_name and display_name != '':
                            self.tables[current_entity]['display_name'] = display_name

    def _parse_attribute(self, row, col_idx) -> Dict[str, Any]:
        """解析单个字段"""
        excel_type = str(row[col_idx.get('数据类型', 5)] or 'String').strip()
        yaml_type = self.TYPE_MAPPING.get(excel_type, 'String')

        field_name = str(row[col_idx.get('字段名称', 2)] or '').strip()
        # 应用命名规则转换字段名
        converted_name = self.naming_engine.convert_attribute_name(field_name)

        # 说明列现在在位置16（新增了两列后）
        description_col = 16

        attr = {
            'name': converted_name,
            'type': yaml_type,
            'display_name': str(row[col_idx.get('显示名称', 3)] or '').strip(),
            'description': str(row[description_col] or '') if len(row) > description_col else '',
            'required': self._to_bool(row[col_idx.get('是否必填', 8)]),
            'is_primary_name': self._to_bool(row[col_idx.get('是否主字段', 11)]),
            'searchable': self._to_bool(row[col_idx.get('可搜索', 12)])
        }

        # 记录原始字段名
        if field_name != converted_name:
            attr['original_name'] = field_name

        # 处理长度
        max_length = row[col_idx.get('长度', 6)]
        if max_length and str(max_length).strip():
            try:
                attr['max_length'] = int(str(max_length).strip())
            except:
                pass

        # 处理默认值
        default_value = row[col_idx.get('默认值', 7)]
        if default_value and str(default_value).strip():
            attr['default_value'] = str(default_value).strip()

        # Lookup字段特殊处理
        if yaml_type in ['Lookup', 'Customer', 'Owner']:
            related_entity = row[col_idx.get('关联实体', 13)]
            if related_entity and str(related_entity).strip():
                # 转换关联实体名
                converted_entity = self.naming_engine.convert_schema_name(str(related_entity).strip())
                attr['entity'] = converted_entity

        # 处理选项集类型（Choice/MultiSelect Choice）
        if yaml_type in ['Picklist', 'MultiSelectPicklist']:
            optionset_ref = str(row[col_idx.get('选项集引用', 14)] or '').strip()
            option_def = str(row[col_idx.get('选项定义', 15)] or '').strip()

            if optionset_ref and option_def:
                # 错误：同时指定了引用和定义
                raise ValueError(f"字段 {field_name} 不能同时引用全局选项集和定义本地选项集")

            if optionset_ref:
                # 引用全局选项集
                attr['option_set_ref'] = self.naming_engine.convert_schema_name(optionset_ref)
            elif option_def:
                # 解析本地选项集定义
                attr['options'] = self._parse_local_options(option_def)
            else:
                # 允许空值，由后续验证处理
                pass

        return attr

    def _parse_local_options(self, option_def: str) -> List[Dict[str, Any]]:
        """解析本地选项集定义，格式：激活=1,禁用=2"""
        options = []
        for part in option_def.split(','):
            part = part.strip()
            if '=' in part:
                label, value = part.split('=', 1)
                try:
                    options.append({
                        'label': label.strip(),
                        'value': int(value.strip())
                    })
                except ValueError:
                    # 如果值不是整数，尝试作为字符串
                    options.append({
                        'label': label.strip(),
                        'value': value.strip()
                    })
        return options

    def _parse_optionsets(self, sheet):
        """解析选项集工作表"""
        current_optionset = None

        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [cell or '' for cell in row]
            break

        col_idx = {name: idx for idx, name in enumerate(headers)}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            optionset_name = str(row[col_idx.get('选项集名称', 0)] or '').strip()

            if optionset_name and optionset_name != '':
                current_optionset = optionset_name
                # 应用命名规则转换选项集名
                schema_name = self.naming_engine.convert_schema_name(optionset_name)

                if current_optionset not in self.optionsets:
                    self.optionsets[current_optionset] = {
                        'original_name': current_optionset,
                        'name': schema_name,
                        'display_name': '',
                        'options': []
                    }

                    display_name = str(row[col_idx.get('显示名称', 3)] or '').strip()
                    if display_name and display_name != '':
                        self.optionsets[current_optionset]['display_name'] = display_name

            # 解析选项
            if current_optionset:
                value_str = str(row[col_idx.get('选项值', 2)] or '').strip()
                if value_str and value_str != '':
                    try:
                        value = int(value_str)
                    except:
                        continue

                    option = {
                        'value': value,
                        'label': str(row[col_idx.get('显示名称', 3)] or '').strip(),
                        'label_en': str(row[col_idx.get('英文显示名称', 4)] or '').strip()
                    }

                    color = row[col_idx.get('颜色标记', 6)]
                    if color and str(color).strip():
                        option['color'] = str(color).strip()

                    self.optionsets[current_optionset]['options'].append(option)

    def _parse_forms(self, sheet):
        """解析表单设计工作表 (04_表单设计)"""
        current_form = None
        current_tab = None
        current_section = None
        row_in_section = 0

        # 从 Row 2 读取表单基本信息
        header_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        form_name = str(header_row[1] or '').strip() if len(header_row) > 1 else ''
        entity_name = str(header_row[4] or '').strip() if len(header_row) > 4 else ''

        if not form_name or not entity_name:
            return

        # 初始化表单
        form_key = f"{entity_name}_main"
        self.forms[form_key] = {
            'original_name': form_name,
            'entity': entity_name,
            'schema_name': f"{entity_name}_main_form",
            'type': 'Main',
            'display_name': form_name,
            'is_default': True,
            'tabs': []
        }

        current_form = self.forms[form_key]

        # 从 Row 3 开始解析布局
        for row in sheet.iter_rows(min_row=3, values_only=True):
            # A 列 (0): Tab 标记或 Section 标记或字段名
            # L 列 (11): 字段 schema name
            # Q 列 (16): Tab 位置标记
            a_col = str(row[0] or '').strip() if len(row) > 0 else ''
            l_col = str(row[11] or '').strip() if len(row) > 11 else ''
            q_col = str(row[16] or '').strip() if len(row) > 16 else ''

            # 检测 Tab 标记: 【Tab N: 名称】
            if a_col.startswith('【Tab') and '】' in a_col:
                # 保存之前的 Section
                if current_section:
                    current_tab['sections'].append(current_section)

                # 保存之前的 Tab
                if current_tab:
                    current_form['tabs'].append(current_tab)

                # 解析新 Tab 名称
                tab_name = a_col.split('】')[0].split(':', 1)[-1].strip()
                tab_index = a_col.split('Tab')[1].split(':')[0].strip() if 'Tab' in a_col else '1'

                current_tab = {
                    'name': f"tab_{self._to_english_name(tab_name)}_{tab_index}",
                    'display_name': tab_name,
                    'expand_by_default': tab_index == '1',
                    'sections': []
                }
                current_section = None
                row_in_section = 0
                continue

            # 检测 Section 标记: 分组: 名称
            if a_col.startswith('分组:'):
                # 保存之前的 Section
                if current_section:
                    current_tab['sections'].append(current_section)

                # 创建新 Section
                section_name = a_col.split(':', 1)[1].strip()
                current_section = {
                    'name': f"sec_{self._to_english_name(section_name)}",
                    'display_name': section_name,
                    'rows': []
                }
                row_in_section = 0
                continue

            # 处理字段
            if l_col and current_tab:
                # 确保有当前 Section
                if not current_section:
                    current_section = {
                        'name': f"sec_default",
                        'display_name': '默认分组',
                        'rows': []
                    }

                # 根据位置判断是单字段还是双字段布局
                # 如果 A 列有内容，可能是字段标签占位
                # 简化处理：每行一个字段，width="2"
                cell = {
                    'attribute': l_col,
                    'width': '2'
                }

                # 检查是否需要新建行
                if not current_section['rows'] or len(current_section['rows'][-1]['cells']) >= 2:
                    current_section['rows'].append({'cells': []})

                current_section['rows'][-1]['cells'].append(cell)
                row_in_section += 1

        # 保存最后的 Section 和 Tab
        if current_section and current_tab:
            current_tab['sections'].append(current_section)
        if current_tab:
            current_form['tabs'].append(current_tab)

    def _parse_views(self, sheet):
        """解析视图定义工作表 (03_视图定义)"""
        current_view = None
        view_key = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            view_name = row[0]  # 视图名称
            entity_name = row[1] if len(row) > 1 else ''  # 实体名称
            field_name = row[3] if len(row) > 3 else ''  # 字段名称

            # 空行跳过
            if not view_name and not field_name:
                continue

            # 新视图开始
            if view_name:
                view_key += 1
                current_view = {
                    'name': view_name,
                    'entity': entity_name or '',
                    'type': 'PublicView',
                    'display_name': view_name,
                    'is_default': False,
                    'columns': []
                }
                self.views[f"{entity_name or 'entity'}_{view_key}"] = current_view

            # 添加列
            if current_view and field_name:
                seq = row[2]  # 序号
                if seq and str(seq).isdigit():
                    col = {
                        'attribute': field_name,
                        'width': self._parse_int(row[5], 100),
                        'sort_order': int(seq),
                        'format': row[10] or 'text',
                        'align': row[6] or 'left',
                        'sortable': self._to_bool(row[7]) if row[7] is not None else True,
                        'filterable': self._to_bool(row[8]) if row[8] is not None else True
                    }
                    current_view['columns'].append(col)

    def _parse_int(self, value: Any, default: int = 0) -> int:
        """解析整数值"""
        try:
            return int(value) if value else default
        except (ValueError, TypeError):
            return default

    def _to_english_name(self, chinese_name: str) -> str:
        """将中文转换为英文逻辑名"""
        # 简单映射常见中文词
        mapping = {
            '常规': 'general',
            '基本信息': 'basic',
            '详细': 'detail',
            '详细信息': 'details',
            '联系': 'contact',
            '联系信息': 'contact_info',
            '公司': 'company',
            '公司信息': 'company_info',
            '地址': 'address',
            '地址信息': 'address_info',
            '关系': 'relation',
            '主要关系': 'primary_relation',
            '财务': 'financial',
            '财务信息': 'financial_info',
            '备注': 'notes',
            '其他': 'other'
        }

        for cn, en in mapping.items():
            if cn in chinese_name:
                return en

        # 如果没有映射，返回拼音首字母或简化版本
        return 'custom'

    def _to_bool(self, value) -> bool:
        """转换为布尔值"""
        if value is None:
            return False
        return str(value).strip() in ['是', 'True', 'true', '1', 'yes', 'Yes']

    def validate_primary_name(self):
        """验证主字段设置"""
        errors = []

        for table_name, table_data in self.tables.items():
            primary_count = sum(1 for attr in table_data['attributes']
                               if attr.get('is_primary_name', False))

            if primary_count == 0:
                errors.append(f"表 {table_name}: 缺少主字段")
            elif primary_count > 1:
                primary_fields = [attr['name'] for attr in table_data['attributes']
                                 if attr.get('is_primary_name', False)]
                errors.append(f"表 {table_name}: 主字段超过一个 ({', '.join(primary_fields)})")

        return errors

    def generate_yaml(self, output_dir: str, include_forms: bool = False, include_views: bool = False) -> Dict[str, Any]:
        """生成YAML文件"""
        generated_files = []
        naming_info = {'tables': {}, 'optionsets': {}, 'forms': {}, 'views': {}}

        tables_dir = Path(output_dir) / 'tables'
        optionsets_dir = Path(output_dir) / 'optionsets'
        forms_dir = Path(output_dir) / 'forms'
        views_dir = Path(output_dir) / 'views'
        tables_dir.mkdir(parents=True, exist_ok=True)
        optionsets_dir.mkdir(parents=True, exist_ok=True)
        if include_forms:
            forms_dir.mkdir(parents=True, exist_ok=True)
        if include_views:
            views_dir.mkdir(parents=True, exist_ok=True)

        # 验证主字段
        validation_errors = self.validate_primary_name()
        if validation_errors:
            print(f"验证警告: {validation_errors}")

        # 生成表YAML
        for table_name, table_data in self.tables.items():
            schema_name = table_data['schema_name']
            yaml_content = {
                '$schema': '../../_schema/table_schema.yaml',
                'schema': table_data
            }

            output_file = tables_dir / f"{schema_name}.yaml"
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(f"# YAML Metadata - Generated from Excel\n")
                f.write(f"# Original Name: {table_name}\n")
                f.write(f"# Naming Rule: lowercase with separator '_'\n\n")
                yaml.dump(yaml_content, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

            generated_files.append(str(output_file))
            naming_info['tables'][table_name] = schema_name
            print(f"生成表文件: {output_file} (原: {table_name} → {schema_name})")

        # 生成选项集YAML
        for optionset_name, optionset_data in self.optionsets.items():
            schema_name = optionset_data['name']
            yaml_content = {
                '$schema': '../../_schema/optionset_schema.yaml',
                'name': schema_name,
                'display_name': optionset_data['display_name'],
                'options': optionset_data['options']
            }

            output_file = optionsets_dir / f"{schema_name}.yaml"
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(f"# YAML Metadata - Option Set\n")
                f.write(f"# Original Name: {optionset_name}\n\n")
                yaml.dump(yaml_content, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

            generated_files.append(str(output_file))
            naming_info['optionsets'][optionset_name] = schema_name
            print(f"生成选项集文件: {output_file} (原: {optionset_name} → {schema_name})")

        # 生成表单YAML
        if include_forms and self.forms:
            for form_key, form_data in self.forms.items():
                yaml_content = {
                    '$schema': '../_schema/form_schema.yaml',
                    'form': {
                        'schema_name': form_data['schema_name'],
                        'entity': form_data['entity'],
                        'type': form_data['type'],
                        'display_name': form_data['display_name'],
                        'is_default': form_data['is_default'],
                        'tabs': form_data['tabs']
                    }
                }

                output_file = forms_dir / f"{form_data['entity']}_main.yaml"
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(f"# YAML Metadata - Form\n")
                    f.write(f"# Original Name: {form_data['original_name']}\n")
                    f.write(f"# Entity: {form_data['entity']}\n\n")
                    yaml.dump(yaml_content, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

                generated_files.append(str(output_file))
                naming_info['forms'][form_key] = form_data['schema_name']
                print(f"生成表单文件: {output_file} (原: {form_data['original_name']})")

        # 生成视图YAML
        if include_views and self.views:
            for view_key, view_data in self.views.items():
                # 构建 fetchxml
                attributes = [col['attribute'] for col in view_data['columns']]
                fetchxml = self._build_fetch_xml(view_data['entity'], attributes)

                yaml_content = {
                    '$schema': '../../_schema/view_schema.yaml',
                    'view': {
                        'schema_name': view_data['name'],
                        'entity': view_data['entity'],
                        'type': view_data['type'],
                        'display_name': view_data['display_name'],
                        'is_default': view_data['is_default'],
                        'fetch_xml': fetchxml
                    },
                    'columns': view_data['columns']
                }

                output_file = views_dir / f"{view_data['entity']}_{view_data['name']}.yaml"
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(f"# YAML Metadata - View\n")
                    f.write(f"# Original Name: {view_data['name']}\n")
                    f.write(f"# Entity: {view_data['entity']}\n\n")
                    yaml.dump(yaml_content, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

                generated_files.append(str(output_file))
                naming_info['views'][view_key] = view_data['name']
                print(f"生成视图文件: {output_file} (原: {view_data['name']})")

        return {
            'files': generated_files,
            'naming_info': naming_info
        }

    def _build_fetch_xml(self, entity: str, attributes: list[str]) -> str:
        """构建 Fetch XML"""
        lines = ['<fetch version="1.0" mapping="logical">']
        lines.append(f'  <entity name="{entity}">')

        for attr in attributes:
            lines.append(f'    <attribute name="{attr}" />')

        lines.append('  </entity>')
        lines.append('</fetch>')

        return "\n".join(lines)


def convert_excel_to_yaml(
    excel_path: str,
    output_dir: str = None,
    config_path: str = None,
    include_forms: bool = False,
    include_views: bool = False
) -> Dict[str, Any]:
    """
    将Excel转换为YAML的主函数

    Args:
        excel_path: Excel文件路径
        output_dir: 输出目录
        config_path: 命名规则配置文件路径
        include_forms: 是否包含表单转换
        include_views: 是否包含视图转换

    Returns:
        转换结果
    """
    if output_dir is None:
        # 自动查找项目根目录
        current_path = Path(excel_path).parent
        while current_path.parent != current_path:
            if (current_path / 'metadata').exists():
                output_dir = str(current_path / 'metadata')
                break
            current_path = current_path.parent
        else:
            output_dir = 'metadata'

    # 查找命名规则配置文件
    if config_path is None:
        current_path = Path(excel_path).parent
        while current_path.parent != current_path:
            config_file = current_path / 'config' / 'naming_rules.yaml'
            if config_file.exists():
                config_path = str(config_file)
                break
            current_path = current_path.parent

    converter = ExcelToYamlConverter(excel_path, config_path)
    success, message = converter.load_excel()

    if not success:
        return {
            'success': False,
            'message': message
        }

    result = converter.generate_yaml(output_dir, include_forms=include_forms, include_views=include_views)

    return {
        'success': True,
        'message': f'转换成功，生成 {len(result["files"])} 个文件',
        'files': result['files'],
        'tables': list(converter.tables.keys()),
        'optionsets': list(converter.optionsets.keys()),
        'forms': list(converter.forms.keys()) if include_forms else [],
        'views': list(converter.views.keys()) if include_views else [],
        'naming_info': result['naming_info']
    }


def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(
        description='将Excel实体设计转换为YAML元数据文件',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  python convert_excel_to_yaml.py design.xlsx
  python convert_excel_to_yaml.py design.xlsx --include-forms
  python convert_excel_to_yaml.py design.xlsx --include-views
  python convert_excel_to_yaml.py design.xlsx --include-forms --include-views
  python convert_excel_to_yaml.py design.xlsx -o metadata/ --include-forms --include-views
        '''
    )
    parser.add_argument('excel_file', help='Excel设计文件路径')
    parser.add_argument('-o', '--output', help='输出目录 (默认: metadata/)')
    parser.add_argument('-c', '--config', help='命名规则配置文件路径')
    parser.add_argument(
        '--include-forms',
        action='store_true',
        help='包含表单转换 (04_表单设计 工作表)'
    )
    parser.add_argument(
        '--include-views',
        action='store_true',
        help='包含视图转换 (03_视图定义 工作表)'
    )
    parser.add_argument(
        '--forms-only',
        action='store_true',
        help='仅转换表单，不转换实体和选项集'
    )
    parser.add_argument(
        '--views-only',
        action='store_true',
        help='仅转换视图，不转换实体和选项集'
    )

    args = parser.parse_args()

    result = convert_excel_to_yaml(
        args.excel_file,
        args.output,
        args.config,
        include_forms=args.include_forms or args.forms_only,
        include_views=args.include_views or args.views_only
    )

    if result['success']:
        print(f"\n{result['message']}")
        if not args.forms_only and not args.views_only:
            print(f"表: {', '.join(result['tables'])}")
            if result['optionsets']:
                print(f"选项集: {', '.join(result['optionsets'])}")
        if result['forms']:
            print(f"表单: {', '.join(result['forms'])}")
        if result['views']:
            print(f"视图: {', '.join(result['views'])}")

        if result.get('naming_info'):
            print("\n命名转换:")
            for orig, converted in result['naming_info']['tables'].items():
                print(f"  表 {orig} → {converted}")
            for orig, converted in result['naming_info']['forms'].items():
                print(f"  表单 {orig} → {converted}")
            for orig, converted in result['naming_info']['views'].items():
                print(f"  视图 {orig} → {converted}")
    else:
        print(f"错误: {result['message']}")
        import sys
        sys.exit(1)


if __name__ == '__main__':
    main()
